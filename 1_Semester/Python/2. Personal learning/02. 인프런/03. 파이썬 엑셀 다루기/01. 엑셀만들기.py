import openpyxl

# 1) 엑셀 만들기  = wb    // 만들고나서 변수에 담기 

wb = openpyxl.Workbook()

# 2) 엑셀 워크시트 만들기  = ws    // 만들고나서 변수에 담기 
ws = wb.create_sheet("우리집")


# 3) 데이터 추가하기
ws['A1'] = '참가번호'
ws['B1'] = '성명'


ws['A2'] = 1
ws['B2'] = '이재일'

# 4) 엑셀 저장하기 
# wb.save('참가자_data.xlsx') 
# 이렇게 앞에 경로를 지정하지 않고 파일이름만 써서 저장하면
# 전체 프로젝트 파일 안에 저장이 된다.
  
# 1. 원하는 파일 찾고
# 2. "경로복사(절대경로)" 후 파일명 앞에 붙여넣기  
# 3. 파일명 앞에 "\" (역슬레쉬) 붙이기 
wb.save(r'C:\Python_File\4. 인프런\03. 파이썬 엑셀 다루기\참가자_data.xlsx')

# #### 문자열 안에서 \는 이스케이프 문자로 인식함 ####
# 그냥 문자로 인식 하게 하기 위해 "r"을 붙인다 => 전부 문자로 인식해라 
# row 문자열로 만들어준다